import re
import openpyxl
import pickle
from openpyxl import Workbook
from text_process import json_content_all_txt, text_process_final


def read_relation_txt(file_path):
    # 从txt文件读取知识三元组，返回关系列表和实体列表
    res_relation = []
    res_entities = []
    with open(file_path, 'r', encoding="utf-8") as read_file:
        for line in read_file.readlines():
            line = re.sub(",Company", ":C", line)
            line = re.sub(",Industry|\n", "", line)
            res_relation.append(line)
            res_entities += relation2entities(line)
    res_relation = list(set(res_relation))
    res_entities = list(set(res_entities))
    print("all relation below:\n")
    print(res_relation)
    print("all entities below:\n")
    print(res_entities)
    return res_relation, res_entities

    return res_relation, res_entities


def read_relation_xlsx(file_path):
    # 从xlsx文件读取知识三元组，返回关系列表和实体列表
    relation_list = []
    entity_lst = []
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    for row in ws.rows:
        for cell in row:
            if cell.value is not None and len(re.split('，|,', cell.value)) >= 3:
                relation_list.append(cell.value.strip())
    relation_list = [re.sub('：', ':', relation) for relation in relation_list]
    relation_list = [re.sub('，', ',', relation) for relation in relation_list]
    # relation_list = [relation.upper() for relation in relation_list]
    relation_list = [re.sub('c', 'C', relation) for relation in relation_list]
    relation_list = [re.sub('l', 'L', relation) for relation in relation_list]

    relation_list = list(set(relation_list))

    for relation in relation_list:
        entity_lst += relation2entities(relation)
    entity_lst = list(set(entity_lst))

    print("all relation below:\n")
    print(relation_list)
    print("all entities below:\n")
    print(entity_lst)
    return relation_list, entity_lst


def read_relation_xlsx_new(file_path):
    # 从xlsx文件读取知识三元组，返回关系列表和实体列表
    relation_list = []
    entity_lst = []
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    for row in ws.rows:
        for cell in row:
            if cell.value is not None and len(re.split('，|,', cell.value)) >= 3:
                cell_list = re.split('，|,', cell.value.strip())
                cell1 = []
                cell2 = []
                if cell_list[0].find("、") or cell_list[0].find("="):
                    cell1 += re.split('、|=', cell_list[0])
                else:
                    cell1 = [cell_list[0]]
                if cell_list[0].find("、") or cell_list[0].find("="):
                    cell2 += re.split('、|=', cell_list[2])
                else:
                    cell2 += [cell_list[2]]
                for c1 in cell1:
                    for c2 in cell2:
                        relation_list.append(c1+','+cell_list[1]+','+c2)

    relation_list = [re.sub('：', ':', relation) for relation in relation_list]
    relation_list = [re.sub('，', ',', relation) for relation in relation_list]
    # relation_list = [relation.upper() for relation in relation_list]
    relation_list = [re.sub('c', 'C', relation) for relation in relation_list]
    relation_list = [re.sub('l', 'L', relation) for relation in relation_list]

    relation_list = list(set(relation_list))

    for relation in relation_list:
        entity_lst += relation2entities(relation)
    entity_lst = list(set(entity_lst))

    print("all relation below:\n")
    print(relation_list)
    print("all entities below:\n")
    print(entity_lst)
    return relation_list, entity_lst


def sentence_has_word(file_path, word1, word2):
    # 判定一个文件中每一行,如果同时包括word1和word2则打印出来
    read_file = open(file_path, encoding="utf-16")
    for line in read_file.readlines():
        if word1 in line and word2 in line:
            print(line)


def find_sentence_in_wordlst(file_path, wordlst, save_path):
    # 从file_path中读取每一行，如果这一行至少含有实体列表中的两个实体，就保存在save_path中
    save_file = open(save_path, 'w', encoding='UTF-8')
    read_file = open(file_path, encoding='UTF-8')
    for line in read_file.readlines():
        pos = 0
        for word in wordlst:
            if word in line:
                pos += 1
        if pos >= 2:
            # print(line)
            save_file.write(line)


def relation2entities(relation):
    # 读取关系relation, 去除关系中可能含有的‘、’‘：C|：c’等字符，返回两个实体
    entity_list = re.split('，|,', relation)
    entities = []
    if len(entity_list) < 3:
        print("关系不完整\n")
        print(relation)
        return []
    if entity_list[0].find("、"):
        entities += entity_list[0].split('、')
    else:
        entities += [entity_list[0]]
    if entity_list[2].find("、"):
        entities += entity_list[2].split('、')
    else:
        entities += [entity_list[2]]

    # p = re.compile(':c|:C|：C|：c|:L|：L')
    entities = list(set(entities))
    # print("关系中的实体有：")
    # print(entities)
    return entities


def sentence_has_relation(sentence, relation):
    # 如果一个句子包含一个关系的两个实体，那么就判定改句子存在该关系
    entities = relation2entities(relation)

    pos = 0
    for word in entities:
        if word in sentence:
            pos += 1
    if pos >= 2:
        return 1
    else:
        return 0


def file_has_relation_list(relation_lst, file_path):
    # 读取文件file_path 和 关系列表 relation_lst，如果每一行能和关系列表的某一关系对应，就打印该行及对应的关系
    # entities = relation2entities(relation)
    # entity_list = re.split('，|,', relation)
    read_file = open(file_path)
    for line in read_file.readlines():
        print(line)
        for relation in relation_lst:

            if sentence_has_relation(line, relation):
                print("存在关系：" + relation)
        print('\n')


def file_find_relation_save(relation_lst, file_path, save_path):
    # 读取文件file_path 和 关系列表 relation_lst，如果每一行能和关系列表的某一关系对应，就保存该行及对应的关系到save_path中
    # entities = relation2entities(relation)
    # entity_list = re.split('，|,', relation)
    save_file = open(save_path, 'w')
    read_file = open(file_path)
    for line in read_file.readlines():
        save_file.write(line)
        for relation in relation_lst:
            if sentence_has_relation(line, relation):
                save_file.write("存在关系：" + relation)
                save_file.write('\n')
        save_file.write('\n')


def file_extract_relation_xlsx(relation_list, file_path, save_xlsx):
    # 读取文件file_path 和 关系列表 relation_lst，如果每一行能和关系列表的某一关系对应，就保存该行及对应的关系到表格中
    book = Workbook()
    sheet = book.active
    read_file = open(file_path, encoding='UTF-8')
    for line in read_file.readlines():
        line_list = [line]
        for relation in relation_list:
            if sentence_has_relation(line, relation):
                line_list.append(relation)
        print(line_list)
        sheet.append(line_list)
    book.save(save_xlsx)


def entity2relation(relation_lst, entity_lst):
    if len(entity_lst)< 2:
        return []
    entity_lst = sorted(entity_lst, key = lambda value:len(value), )
    exist_relationlst = []
    p = re.compile(':c|:C|：C|：c|:L|：L')

    # 对互相包含的实体消除重复，比如文本中同时出现锂电池、锂、电池，那么只保留锂电池
    remove_lst = []
    for i in range(0, len(entity_lst) - 1):
        for j in range(i + 1, len(entity_lst)):
            if re.sub(p, '', entity_lst[i]) in re.sub(p, '', entity_lst[j]):
                remove_lst.append(entity_lst[i])
    remove_lst = list(set(remove_lst))
    for value in remove_lst:
        entity_lst.remove(value)

    for i in range(0, len(entity_lst) - 1):
        for j in range(i + 1, len(entity_lst)):
            pos = 0
            for relation in relation_lst:
                if entity_lst[i] in relation and entity_lst[j] in relation:
                    exist_relationlst.append(relation)
                    pos = 1
                    break
            if pos == 0:
                exist_relationlst.append(entity_lst[i] + ",共现," + entity_lst[j])
    return list(set(exist_relationlst))


def entity2relation2(relation_lst, entity_lst):
    exist_relationlst = []
    for i in range(0, len(entity_lst) - 1):
        for j in range(i + 1, len(entity_lst)):
            for relation in relation_lst:
                if entity_lst[i] in relation and entity_lst[j] in relation:
                    exist_relationlst.append(relation)
    return list(set(exist_relationlst))


def file_extract_all_relation_xlsx(relation_list, entity_list, file_path, save_xlsx, none_relation_max, have_relation_max):
    # 读取文件file_path 和 关系列表 relation_lst，如果每一行能和关系列表的某一关系对应，就保存该行及对应的关系到表格中
    book = Workbook()
    sheet = book.active
    read_file = open(file_path, encoding='UTF-8')
    p = re.compile(':c|:C|：C|：c|:L|：L')
    none_num = 0
    have_num = 0
    for line in read_file.readlines():
        line_list = [line]
        exist_entitylst = [entity for entity in entity_list if re.sub(p, '', entity) in line]
        exist_entitylst = list(set(exist_entitylst))
        exist_relationlst = entity2relation(relation_list, exist_entitylst)
        if len(exist_relationlst) > 0 and have_num < have_relation_max:
            exist_relationlst = entity2relation(relation_list, exist_entitylst)
            have_num += 1
            # print("存在的实体： ")
            # print(exist_entitylst)
            # print("存在的关系： ")
            # print(exist_relationlst)
            line_list += exist_relationlst
            sheet.append(line_list)
        elif len(exist_relationlst) == 0 and none_num < none_relation_max:
            none_num += 1
            sheet.append(line_list)
        if none_num == none_relation_max and have_num == have_relation_max:
            break
    book.save(save_xlsx)


def save_relation_xlsx(relation_list, save_xlsx):
    # 将关系列表保存在表格中
    book = Workbook()
    sheet = book.active
    result_list = [relation_list[i:i + 10] for i in range(0, len(relation_lst), 10)]
    for result in result_list:
        sheet.append(result)
    book.save(save_xlsx)


def relation_del_same(relation_lst, relation_name_lst):
    # 输入关系列表，删除其中存在的重复关系
    del_num = 0
    for relation in relation_lst:
        entity_list = re.split('，|,', relation)
        first_string = entity_list[1]
        if first_string in relation_name_lst:
            # print(relation)
            second_string = re.sub(first_string, '子行业', relation)
            if second_string in relation_lst:
                print(second_string + "\t 已经被删除")
                relation_lst.remove(second_string)
                del_num += 1
    # print(relation_lst)
    print("\n共删除 " + str(del_num) + ' 条关系！\n')
    return relation_lst


def add_coexist_relation(relation_lst):
    # 增加非直接相关的两个实体间的共现关系到关系列表中
    entities = []
    replace_lst = []
    for relation in relation_lst:
        entity_list = re.split('，|,', relation)
        entities += [entity_list[0]]
        entities += [entity_list[2]]
        temp_string1 = entity_list[0] + ",共现," + entity_list[2]
        temp_string2 = entity_list[2] + ",共现," + entity_list[0]
        replace_lst.append(temp_string1)
        replace_lst.append(temp_string2)
    entities = [entity.upper() for entity in entities]
    entities = list(set(entities))
    # 注意此处返回的实体列表，实体保留：C、：L，仅供此处构建共现关系使用；
    # 不保留此符号的实体列表从《read_relation_xlsx》函数返回；
    print("共现关系中的实体有：")
    print(entities)

    coexist_lst = []
    print("关系列表的长度是： " + str(len(relation_lst)))
    print("实体列表的长度是： " + str(len(entities)))
    for i in range(0, len(entities) - 1):
        for j in range(i, len(entities) - 1):
            temp_string = entities[i] + ",共现," + entities[j]
            coexist_lst.append(temp_string)
    print("共现关系列表构建完成！")
    print("共现列表的长度：" + str(len(coexist_lst)))
    # print(coexist_lst)
    del_num = 0
    for replace_relation in replace_lst:
        if replace_relation in coexist_lst:
            coexist_lst.remove(replace_relation)
            del_num += 1

    print("共删除" + str(del_num) + "条关系！")
    # print("删除后的共享列表：")
    # print(coexist_lst)
    return list(set(coexist_lst + relation_lst))


def save_pickle(data, data_path):
    # 保存数据到pickle中
    with open(data_path, 'wb') as f:
        # Pickle the 'data' dictionary using the highest protocol available.
        pickle.dump(data, f)


def load_pickle(save_path):
    # 从pickle中读取数据
    with open(save_path, 'rb') as f:
        # The protocol version used is detected automatically, so we do not
        # have to specify it.
        data = pickle.load(f)
    return data


def find_relation(relation_path, corpus_path, save_xlsx_path, none_max, have_max):
    # result_relation, result_entity = read_relation_xlsx_new(relation_path)
    result_relation, result_entity = read_relation_txt(relation_path)
    if '行业' in result_entity:
        result_entity.remove('行业')
    result_relation = relation_del_same(result_relation, ['下游'])
    file_extract_all_relation_xlsx(result_relation, result_entity, corpus_path, save_xlsx_path, none_max, have_max)


if __name__ == "__main__":
    relation_path = "data/entity_relations.txt"
    semi_json_path = "data/semi/semi_Tsh_0726.json"
    semi_corpus_all = "data/semi/semi_corpus_tsh_all.txt"
    semi_corpus_tsh = "data/semi/semi_tsh_short3.txt"
    semi_tsh_save_xlsx = "data/semi/semi_corpus_tsh.xlsx"

    new_car_relation_path = "data/new_car/新能源汽车_0715_mod.xlsx"
    new_car_json_path = "data/new_car/new_car_Tsh_0725.json"
    new_car_file_all = "data/new_car/new_car_chyxx_result.txt"
    new_car_corpus_tsh = "data/new_car/new_car_tsh_short3.txt"
    new_car_tsh_save_xlsx = "data/new_car/new_car_corpus_tsh.xlsx"

    # json_content_all_txt(semi_json_path, semi_corpus_all)
    # text_process_final(semi_corpus_all, semi_corpus_tsh, 80, 100)
    find_relation(relation_path, semi_corpus_tsh, semi_tsh_save_xlsx, 50, 50)


